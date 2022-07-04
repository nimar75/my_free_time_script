import boto3
import openpyxl
import yaml

client = boto3.client('iam')
# enter the path to you excel file, if you have csv,convert it to excel.
path =
my_file = openpyxl.load_workbook(path)
sheet_obj = my_file.active
row = sheet_obj.max_row
colum = sheet_obj.max_column
x = 0
y = 0

for i in range(1, colum+1, 1):
    if sheet_obj.cell(row=1, column=i).value == "access_key_1_active":
        x = i
    if sheet_obj.cell(row=1, column=i).value == "user":
        y = i
for j in range(1, row+1, 1):
    if sheet_obj.cell(row=j, column=x).value == "true":
        print(sheet_obj.cell(row=j, column=y).value)
